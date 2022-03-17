import time

from fpdf import FPDF
from openpyxl import load_workbook
from selenium import webdriver

List =[]
# code for opening the url and login
driver = webdriver.Chrome(executable_path='C:/Users/crochet/PycharmProjects/First-test/Chrome/chromedriver11.exe')
driver.maximize_window()
filepath="test-login.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
b1 = sheet['B2']
driver.get(b1.value)
List.append("Url opened successfully")
em=driver.find_element_by_xpath("//input[@id='email']")
c2=sheet['C2']
em.send_keys(c2.value)
List.append("UN entered successfully")
em=driver.find_element_by_xpath("//input[@id='password']")
d2=sheet['D2']
em.send_keys(d2.value)
List.append("password entered successfully")
bt=driver.find_element_by_xpath("//div[@class='text-center login_btn']/button[@class='btn btn-primary']").click()
List.append("Signin button clicked")
time.sleep(4)
Title = driver.find_element_by_class_name('navbar-brand').is_displayed()
if Title==True:
    sheet['E2'] = "Yes"
    #alternate way
    #sheet.cell(row=2, column=5).value = 'Yes'
else:
    sheet.cell(row=2, column=5).value = 'No'
wb.save('test-login.xlsx')
pdf = FPDF()
pdf.add_page()
pdf.set_font("Arial", size=15)
pdf.cell(200, 10,txt="Steps",ln=1, align='C')
for i in range(len(List)):
   pdf.cell(200, 10,txt=str(List[i]),ln=2, align='C')
pdf.output("GFG.pdf")

import yagmail
receiver = "manisha.1wayit@outlook.com"
body = "Hello there from testing"
filename = "GFG.pdf"
yag = yagmail.SMTP("manisha.1wayit@gmail.com","Welcome@1way")
yag.send(to=receiver,subject="Yagmail test with attachment",contents=body,attachments=filename)


# email code with pdf
# body = 'Hello Email sending code'
# sender = 'manisha.1wayit@gmail.com'
# password = 'Welcome@1way'
# receiver = 'manisha.1wayit@outlook.com'
# #Setup the MIME
# message = MIMEMultipart()
# message['From'] = sender
# message['To'] = receiver
# message['Subject'] = 'This email has an attacment, a pdf file'
# message.attach(MIMEText(body, 'plain'))
# pdfname = 'GFG.pdf'
# binary_pdf = open(pdfname, 'rb')
# payload = MIMEBase('application', 'octate-stream', Name=pdfname)
# # payload = MIMEBase('application', 'pdf', Name=pdfname)
# payload.set_payload((binary_pdf).read())
# # enconding the binary into base64
# encoders.encode_base64(payload)
# # add header with pdf name
# payload.add_header('Content-Decomposition', 'attachment', filename=pdfname)
# message.attach(payload)
# #use gmail with port
# session = smtplib.SMTP('smtp.gmail.com', 587)
# #enable security
# session.starttls()
# #login with mail_id and password
# session.login(sender, password)
# text = message.as_string()
# session.sendmail(sender, receiver, text)
# session.quit()
# print('Mail Sent')

