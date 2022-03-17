import unittest
from openpyxl import load_workbook

class Test(unittest.TestCase):

    def test_read_excel_file(self):

        # set file path
        filepath="test-login.xlsx"
        # load excel-example.xlsx
        wb=load_workbook(filepath)
        # activate demo.xlsx
        sheet=wb.active
        # get b1 cell value
        b1=sheet['B2']
        print(b1.value)
        print("b2 --> ", b1.value)
        # get b2 cell value
        c2=sheet['C2']
        print("c2 --> ", c2.value)
        d2 = sheet['D2']
        print("D2 --> ", d2.value)
        # get b3 cell value
        d3=sheet.cell(row=3,column=4)
        print("b3 --> ", d3.value)

        if jk.value == 1:
            print(jk.value)
        else:
            print(jk.value)
            print("noooooo")

